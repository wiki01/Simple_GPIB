import sys
from PyQt5.QtWidgets import QTextEdit, QInputDialog, QWidget, QTabWidget, QVBoxLayout, QFrame, QLabel, QTableWidget, QPushButton, QMessageBox, QTableWidgetItem, QApplication
from PyQt5 import QtCore
from PyQt5.QtCore import QTimer
from qtstyles import StylePicker
StylePicker().available_styles
import visa
from datetime import datetime
import time
import os
from openpyxl import Workbook
from openpyxl.styles import Font





class Text_Gpib(QTextEdit):
    def __init__(self, parent):
        super().__init__(parent)
    def mousePressEvent(self, event):

        text, press = QInputDialog.getInt(self, 'GPIB 번호입력', 'GPIB 주소를 입력해주세요.')
        if press:
            self.setText(str(text))

class Text_Cycle(QTextEdit):
    def __init__(self, parent):
        super().__init__(parent)
    def mousePressEvent(self, event):

        text, press = QInputDialog.getInt(self, '주기 입력', '주기를 입력해 주세요.')
        if press:
            self.setText(str(text))

class Text_Command(QTextEdit):
    def __init__(self, parent):
        super().__init__(parent)
    def mousePressEvent(self, event):

        text, press = QInputDialog.getMultiLineText(self, '명령어 입력', '명령어를 입력해 주세요.')

        # 텍스트 공백제거
        List = text.split('\n')
        List_Command = []
        for i in range(0, len(List)):
            if List[i] != '':
                List_Command.append('{}{}'.format(List[i], '\n'))

        text = ''.join(List_Command)
        print(text)
        if press:
            self.setText(str(text))


class Text_Count(QTextEdit):
    def __init__(self, parent):
        super().__init__(parent)

    def mousePressEvent(self, event):
        text, press = QInputDialog.getInt(self, '반복 횟수 입력', '반복할 횟수를 입력해 주세요.')
        if press:
            self.setText(str(text))

class MainFrame(QWidget):

    def __init__(self):
        super().__init__()

        # 필요한 Tab 위젯 생성
        self.Tab_Widget = QTabWidget()
        self.setWindowTitle('GPIB_Commuication')
        self.setGeometry(0, 0, 650, 390)
        self.Tab_Widget.addTab(DataFrame(), 'Instrument')

        self.vbox = QVBoxLayout()
        self.vbox.addWidget(self.Tab_Widget)
        self.setLayout(self.vbox)
        self.show()

class DataFrame(QWidget):

    def __init__(self):
        super().__init__()


        # 필요한 Tab 위젯 생성
        self.QFrame = QFrame(self)
        self.Line_Gpib = QLabel(self.QFrame)
        self.Line_Stuts = QLabel(self.QFrame)
        self.Line_Command = QLabel(self.QFrame)
        self.Line_Cycle = QLabel(self.QFrame)
        self.Line_Count = QLabel(self.QFrame)

        self.Text_Gpib = Text_Gpib(self.QFrame)
        self.Text_Stuts = QTextEdit(self.QFrame)
        self.Text_Command = Text_Command(self.QFrame)
        self.Text_Cycle = Text_Cycle(self.QFrame)
        self.Text_Count = Text_Count(self.QFrame)

        self.Table_Data = QTableWidget(self.QFrame)
        self.Btn_Search = QPushButton(self.QFrame)
        self.Btn_Run = QPushButton(self.QFrame)
        self.Btn_Stop = QPushButton(self.QFrame)
        self.Btn_toExcel = QPushButton(self.QFrame)
        self.Btn_Table_Clear = QPushButton(self.QFrame)
        self.Btn_Timeout_Hide = QPushButton(self.QFrame)

        # 명령어 변수 지정
        self.Run = 'RUN'
        self.Stop = 'STOP'
        self.IDN = '*IDN?'
        self.PC = 0
        self.Timeout = 500

        #-----------------Position-------------
        self.QFrame.setGeometry(0, 0, 630, 310)
        self.QFrame.setFrameStyle(QFrame.Box | QFrame.Raised)
        self.QFrame.setLineWidth(2)
        self.QFrame.setMidLineWidth(3)

        self.Line_Gpib.setText('GPIB      : ')
        self.Line_Gpib.setGeometry(10, 5, 70, 25)
        self.Text_Gpib.setGeometry(85, 5, 70, 25)

        self.Line_Stuts.setText('상태       : ')
        self.Line_Stuts.setGeometry(10, 5 + 20 + 10, 70, 25)
        self.Text_Stuts.setGeometry(85, 5 + 20 + 10, 70, 25)
        self.Text_Stuts.setReadOnly(True)

        self.Line_Command.setText('명령어    : ')
        self.Line_Command.setGeometry(10, 5 + 40 + 20, 70, 25)
        self.Text_Command.setGeometry(85, 5 + 40 +20, 70, 25)

        self.Line_Cycle.setText('주기(초)  : ')
        self.Line_Cycle.setGeometry(10, 5 + 60 + 30, 70, 25)
        self.Text_Cycle.setGeometry(85, 5 + 60 + 30, 70, 25)

        self.Line_Count.setText('반복횟수 : ')
        self.Line_Count.setGeometry(10, 5 + 80 + 40, 70, 25)
        self.Text_Count.setGeometry(85, 5 + 80 + 40, 70, 25)

        self.Table_Data.setGeometry(170, 10, 370, 290)
        self.Table_Data.setRowCount(1000)
        self.Table_Data.setColumnCount(3)
        self.Table_Data.setHorizontalHeaderLabels(['명령어', '일시', '결과'])
        for i in range(0, 3):
            self.Table_Data.setColumnWidth(i, 100)

        self.Btn_Search.setGeometry(550, 10, 60, 40)
        self.Btn_Search.setText('Search')
        self.Btn_Search.clicked.connect(self.Searching)
        self.Btn_Search.installEventFilter(self)
        self.Btn_Run.setGeometry(550, 60, 60, 40)
        self.Btn_Run.setText('RUN')
        self.Btn_Run.clicked.connect(self.Run_Command)
        self.Btn_Run.installEventFilter(self)

        self.Btn_Stop.setGeometry(550, 110, 60, 40)
        self.Btn_Stop.setText('STOP')
        self.Btn_Stop.clicked.connect(self.Stoping)
        self.Btn_Stop.installEventFilter(self)

        self.Btn_toExcel.setGeometry(550, 160, 60, 40)
        self.Btn_toExcel.setText('Excel')
        self.Btn_toExcel.clicked.connect(self.toExcel)


        self.Btn_Table_Clear.setGeometry(550, 210, 60, 40)
        self.Btn_Table_Clear.setText('Delete')
        self.Btn_Table_Clear.clicked.connect(self.Table_Data2Clear)

        self.Btn_Timeout_Hide.setGeometry(550, 260, 60, 40)
        self.Btn_Timeout_Hide.setText('Timing')
        self.Btn_Timeout_Hide.clicked.connect(self.Timeout_Setting)

    def Timeout_Setting(self):
        text, press = QInputDialog.getInt(self, '장비 타임아웃 설정', '현재 {} 로 타입아웃이 설정되어 있습니다'.format(self.Timeout))
        if press:
            self.Timeout = int(text)

    def Table_Data2Clear(self):

        ans = QMessageBox.question(self, '기록 삭제', '기록을 삭제하시겠습니다?', QMessageBox.Yes, QMessageBox.No)
        if ans == QMessageBox.Yes:
            self.Table_Data.clear()

    def Searching(self):
        global DUT
        rm = visa.ResourceManager()
        List_Gpib = []
        Index = Main.Tab_Widget.currentIndex()
        Text = self.Text_Gpib.toPlainText()
        try:
            if int(Text) > 1:
                DUT = rm.open_resource('GPIB::{}'.format(int(Text)))
                DUT.write(self.IDN)
                Model_Name = DUT.read()
                Main.Tab_Widget.setTabText(Index, '{}-{}'.format(Model_Name.split(",")[0], Model_Name.split(",")[1]))

                self.Text_Stuts.setText('Connect')

        except:
            res = ['Yes', 'No']
            ans, press = QInputDialog.getItem(self, 'GPIB 자동연결', '{} 해당번호는 연결되지 않음\n자동연결을 실행할까요?'.format(Text), res, 0, False)
            if press:
                if ans == 'Yes':
                    for i in range(1, 20):
                        try:
                            DUT = rm.open_resource('GPIB::{}'.format(i))
                            DUT.write(self.IDN)
                            Model_Name = DUT.read()
                            Main.Tab_Widget.setTabText(Index, '{}-{}'.format(Model_Name.split(",")[0],
                                                                             Model_Name.split(",")[1]))
                            if len(Model_Name) > 0:
                                List_Gpib.append('GPIB::{}'.format(i))
                            break

                        except:
                            pass
                    if len(List_Gpib) == 0:  # GPIB 연결 안됨
                        QMessageBox.information(self, 'GPIB 연결 없음', 'GPIB 연결상태를 확인 후 실행해 주세요.')
                        return
                    self.Text_Gpib.setText(str(int(List_Gpib[0].split('::')[1])))
                    self.Text_Stuts.setText('Connect')

    def Run_Command(self):
        global List_Command
        # 연결상태 확인 및 내용 확인
        try:
            if DUT:
                DUT.timeout = self.Timeout
                pass
        except:
            QMessageBox.information(self, 'GPIB 연결 없음', 'GPIB 연결상태를 확인 후 실행해 주세요.')
            return

        if self.Text_Stuts.toPlainText() != 'Connect':
            QMessageBox.information(self, 'GPIB 연결 없음', 'GPIB 연결상태를 확인 후 실행해 주세요.')
            return

        if self.Text_Command.toPlainText() == '':
            QMessageBox.information(self, '명령어 없음', '명령어를 입력 후 실행해 주세요.')
            return

        if self.Text_Cycle.toPlainText() == '':
            QMessageBox.information(self, '주기 없음', '주기를 입력 후 실행해 주세요.')
            return

        if self.Text_Count.toPlainText() == '':
            QMessageBox.information(self, '반복횟수 없음', '반복횟수를 입력 후 실행해 주세요.')
            return

        List = self.Text_Command.toPlainText()
        List = List.split('\n')

        # 리스트 공백 제거
        List_Command = []
        for i in range(0, len(List)):
            if List[i] != '':
                List_Command.append('{}{}'.format(List[i], '\n'))

        text = ''.join(List_Command)
        self.Text_Command.setText(text)


        self.Timer = QTimer()
        self.Timer.timeout.connect(self.Runing)
        self.Timer.start(int(self.Text_Cycle.toPlainText())*1000)

    def Runing(self):

        # 실행 및 기록
        Count_Stack = 0
        while str(type(self.Table_Data.item(0 + Count_Stack, 0))) != "<class 'NoneType'>":
            Count_Stack = Count_Stack + 1

        for j in range(0, len(List_Command)):
            Run_Command = List_Command[j].replace('\n', '')

            # 기록하고 시작
            Item_Command = QTableWidgetItem()
            Item_Command.setText(Run_Command)

            Item_Data = QTableWidgetItem()
            Item_Data.setText('{}일 {}시 {}분 {}초'.format(datetime.today().day, datetime.today().hour,
                                                           datetime.today().minute, datetime.today().second))

            self.Table_Data.setItem(Count_Stack, 0, Item_Command)
            self.Table_Data.setItem(Count_Stack, 1, Item_Data)

            try:
                DUT.write(Run_Command)
            except:
                pass

            try:
                Reading = DUT.read()
                Item_Result = QTableWidgetItem()
                Item_Result.setText(Reading)
                self.Table_Data.setItem(Count_Stack, 2, Item_Result)
            except:
                Item_Result = QTableWidgetItem()
                Item_Result.setText('None')
                self.Table_Data.setItem(Count_Stack, 2, Item_Result)

            Count_Stack = Count_Stack + 1

            # 주기 만큼 정지
            time.sleep(int(self.Text_Cycle.toPlainText()))
            self.Table_Data.resizeColumnsToContents()
            self.PC = self.PC + 1
        if self.PC >= int(self.Text_Count.toPlainText()):
            self.Timer.stop()
            QMessageBox.information(self, '측정 종료', '측정이 종료되었습니다')

    def Stoping(self):
        self.Timer.stop()
        QMessageBox.information(self, '측정 종료', '측정이 종료되었습니다')


    def toExcel(self):
        # 바탕화면 주소

        Path_desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
        ans = QMessageBox.Yes
        if os.path.exists(Path_desktop + '\\GPIB_Result.xlsx'):
            ans = QMessageBox.question(self, '파일 확인', '파일이 이미 존재 합니다 Yes 를 누르면 변경됩니다.', QMessageBox.Yes, QMessageBox.No)
        if ans == QMessageBox.Yes:
            wb = Workbook()
            ws = wb.active
            Model_Name = Main.Tab_Widget.currentIndex()
            ws.title = Main.Tab_Widget.tabText(Model_Name)
            ws['A1'] = '명령어'
            ws['B1'] = '일시'
            ws['C1'] = '결과'

            a1 = ws['A1']
            b1 = ws['B1']
            c1 = ws['C1']
            ft = Font(name='Arial', size=14, bold=True)
            a1.font = ft
            b1.font = ft
            c1.font = ft

            row = 0
            try:
                while self.Table_Data.item(row, 0).text() !='':
                    Item_Command = self.Table_Data.item(row, 0).text()
                    Item_Data = self.Table_Data.item(row, 1).text()
                    Item_Result = self.Table_Data.item(row, 2).text()

                    ws['A{}'.format(row + 2)].value = Item_Command
                    ws['B{}'.format(row + 2)].value = Item_Data
                    ws['C{}'.format(row + 2)].value = Item_Result
                    row = row + 1
            except:
                wb.save(Path_desktop + '\\GPIB_Result.xlsx')
                QMessageBox.information(self, '생성완료', Path_desktop + '\\GPIB_Result.xlsx 이 생성되었습니다')

    def eventFilter(self, obj, event):

        if event.type() == QtCore.QEvent.MouseButtonPress:
            if event.button() == QtCore.Qt.RightButton:
                if obj.text() == self.Run:
                    Model_Name = Main.Tab_Widget.currentIndex()
                    Model_Name = Main.Tab_Widget.tabText(Model_Name)
                    text, press = QInputDialog.getText(self, '동작 명령어 설정', '{}의 동작 명령어를 입력해주세요'.format(
                        Model_Name
                    ))
                    if press:
                        self.Run = text
                        obj.setText(text)
                elif obj.text() == self.Stop:
                    Model_Name = Main.Tab_Widget.currentIndex()
                    Model_Name = Main.Tab_Widget.tabText(Model_Name)
                    text, press = QInputDialog.getText(self, '정지 명령어 설정', '{}의 정지 명령어를 입력해주세요'.format(
                        Model_Name
                    ))
                    if press:
                        self.Stop = text
                        obj.setText(text)
                elif obj.text() == 'Search':
                    text, press = QInputDialog.getText(self, 'IDN 확인 명령어 입력', '{} 의 확인 명령어를 입력해주세요'.format(
                        self.IDN
                    ))
                    if press:
                        self.IDN = text
                        print(self.IDN)

        return QtCore.QObject.event(obj, event)


if __name__=='__main__':
    app = QApplication(sys.argv)
    Main = MainFrame()
    Main.setStyleSheet(StylePicker("qdark").get_sheet())
    sys.exit(app.exec_())